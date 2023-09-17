import re
from openpyxl import Workbook

# Define the regular expression patterns for size and color
size_pattern = "(\d+'?\d*\"?)\s*(?:x\s*(\d+'?\"?))?"
supported_colors = "(Black|Red|Yellow|Blue|Olive|Gold|Ivory|Green|Beige|Grey|Brown|Multi|Sage|Peacock)"
color_pattern = supported_colors + "([&/, ]+" + supported_colors + ")*"

# Create a new workbook and select the active worksheet
workbook = Workbook()
worksheet = workbook.active

# Write the headers for the size and color columns
worksheet.cell(row=1, column=1, value="Size")
worksheet.cell(row=1, column=2, value="Color")

# Open the file and read each line
with open("data.txt") as f:
    # Initialize a row counter
    row = 2
    for line in f:
        # Strip any whitespace from the line
        line = line.strip()
        # Search for the size and color matches in the line
        size_match = re.search(size_pattern, line)
        color_match = re.search(color_pattern, line)
        # If either match is found, write the size or color to the worksheet
        if size_match or color_match:
            # Initialize an empty list to store the output
            output = []
            # If size match is found, get the size groups from the match object and append to the output list
            if size_match:
                size1 = size_match.group(1)
                size2 = size_match.group(2)
                size_string = size1
                if size2:
                    size_string += " x " + size2
                output.append(size_string)
            # If color match is found, get the color string from the match object and append to the output list
            if color_match:
                color_string = color_match.group(0)
                output.append(color_string)
            # Write the output list to the worksheet
            worksheet.cell(row=row, column=1, value=output[0])
            if len(output) > 1:
                worksheet.cell(row=row, column=2, value=output[1])
            # Increment the row counter
            row += 1

# Save the workbook to a file
workbook.save("result.xlsx")
    
# Wait for user input to exit
input("Press Enter to exit...")